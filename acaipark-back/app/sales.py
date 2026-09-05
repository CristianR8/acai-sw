from __future__ import annotations

from collections import defaultdict
import re
import unicodedata

from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy import case, func
from sqlalchemy.orm import Session

from . import db, models, schemas

router = APIRouter(prefix="/sales", tags=["sales"])
COLOMBIA_TZ = ZoneInfo("America/Bogota")


def _period_start(period: str | None) -> datetime | None:
    if not period:
        return None
    if period == "all":
        return None
    days_by_period = {
        "week": 7,
        "month": 30,
        "quarter": 90,
        "year": 365,
    }
    days = days_by_period.get(period)
    if days is None:
        raise HTTPException(
            status_code=400,
            detail="Periodo invalido. Usa: all, week, month, quarter, year",
        )
    return datetime.now(timezone.utc) - timedelta(days=days)


@router.get("", response_model=list[schemas.SaleOut])
def list_sales(period: str | None = None, db_session: Session = Depends(db.get_db)):
    query = db_session.query(models.Sale)
    start_date = _period_start(period)
    if start_date is not None:
        query = query.filter(models.Sale.created_at >= start_date)
    return query.order_by(models.Sale.id.desc()).limit(200).all()


@router.get("/{sale_id}", response_model=schemas.SaleOut)
def get_sale(sale_id: int, db_session: Session = Depends(db.get_db)):
    sale = db_session.query(models.Sale).filter(models.Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    return sale


def _daily_payment_summary(day: date, db_session: Session) -> dict[str, Decimal]:
    """Return one Colombia calendar day, grouping legacy card payments as dataphone."""
    start = datetime.combine(day, time.min, tzinfo=COLOMBIA_TZ)
    end = start + timedelta(days=1)
    rows = (
        db_session.query(
            models.Sale.payment_method,
            func.coalesce(func.sum(models.Sale.total), 0).label("total"),
        )
        .filter(models.Sale.created_at >= start, models.Sale.created_at < end)
        .group_by(models.Sale.payment_method)
        .all()
    )

    totals: dict[str, Decimal] = {
        "cash": Decimal("0"),
        "transfer": Decimal("0"),
        "dataphone": Decimal("0"),
    }
    for row in rows:
        method = (row.payment_method or "cash").strip().lower()
        key = "dataphone" if method in {"card", "dataphone"} else method
        if key in totals:
            totals[key] += Decimal(row.total or 0)
    totals["total"] = sum(totals.values(), Decimal("0"))
    return totals


@router.get(
    "/summary/daily-payment-methods",
    response_model=schemas.DailyPaymentMethodSummaryOut,
)
def daily_payment_methods(
    day: date,
    db_session: Session = Depends(db.get_db),
):
    totals = _daily_payment_summary(day, db_session)

    return schemas.DailyPaymentMethodSummaryOut(
        date=day,
        cash_total=totals["cash"],
        transfer_total=totals["transfer"],
        dataphone_total=totals["dataphone"],
        total=totals["total"],
    )


@router.get("/summary/daily-payment-methods.xlsx")
def export_daily_payment_methods_xlsx(
    day: date,
    db_session: Session = Depends(db.get_db),
):
    try:
        import openpyxl  # type: ignore
    except Exception:
        raise HTTPException(
            status_code=501,
            detail="Exportar a Excel requiere `openpyxl` instalado en el backend",
        )

    totals = _daily_payment_summary(day, db_session)
    daily_expenses = (
        db_session.query(models.FixedExpensePayment)
        .join(models.FixedExpense)
        .filter(
            models.FixedExpensePayment.status == "manual",
            models.FixedExpensePayment.due_date == day,
        )
        .order_by(models.FixedExpense.name.asc())
        .all()
    )
    total_expenses = sum(
        (Decimal(expense.amount or 0) for expense in daily_expenses),
        Decimal("0"),
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Cierre de caja"
    styles = openpyxl.styles
    dark_green = "174D3D"
    light_green = "E8F1ED"
    pale_green = "DFF0D8"
    border = styles.Border(
        left=styles.Side(style="thin", color="B7C3BE"),
        right=styles.Side(style="thin", color="B7C3BE"),
        top=styles.Side(style="thin", color="B7C3BE"),
        bottom=styles.Side(style="thin", color="B7C3BE"),
    )

    def section(title: str, row: int):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = sheet.cell(row=row, column=1, value=title)
        cell.fill = styles.PatternFill("solid", fgColor=dark_green)
        cell.font = styles.Font(bold=True, color="FFFFFF", size=13)
        cell.alignment = styles.Alignment(horizontal="left")

    def header(row: int, values: list[str]):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.fill = styles.PatternFill("solid", fgColor=light_green)
            cell.font = styles.Font(bold=True, color=dark_green, size=12)
            cell.alignment = styles.Alignment(horizontal="center")
            cell.border = border

    def money_row(row: int, label: str, reference: str, amount: Decimal, *, total: bool = False):
        values = [label, reference, float(amount)]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = border
            cell.alignment = styles.Alignment(
                horizontal="right" if column == 3 else "left",
                vertical="center",
            )
            if total:
                cell.fill = styles.PatternFill("solid", fgColor=light_green)
                cell.font = styles.Font(bold=True, size=12)
        sheet.cell(row=row, column=2).font = styles.Font(italic=True, color="6B7280") if not total else styles.Font(bold=True, size=12)
        sheet.cell(row=row, column=3).number_format = '$#,##0'

    section("3. RESUMEN DE VENTAS Y MEDIOS DE PAGO", 1)
    sheet.cell(row=2, column=1, value="Fecha del cierre")
    sheet.cell(row=2, column=2, value=day.strftime("%d/%m/%Y"))
    header(4, ["Medio de Pago", "Comprobante / Ref", "Monto Sistema ($)"])
    money_row(5, "Ventas en Efectivo", "POS Sistema", totals["cash"])
    money_row(6, "Datáfono / Tarjetas", "Vouchers / Lote", totals["dataphone"])
    money_row(7, "Transferencias", "Transferencias", totals["transfer"])
    money_row(8, "TOTAL VENTAS REGISTRADAS", "", totals["total"], total=True)

    expense_start = 10
    section("4. GASTOS REGISTRADOS DEL DÍA", expense_start)
    header(expense_start + 1, ["Gasto", "Categoría / Ref", "Monto Sistema ($)"])
    row = expense_start + 2
    if daily_expenses:
        for expense in daily_expenses:
            money_row(
                row,
                expense.fixed_expense.name,
                expense.fixed_expense.category or "Gasto manual",
                Decimal(expense.amount or 0),
            )
            row += 1
    else:
        for column, value in enumerate(["Sin gastos registrados", "", 0], start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = border
        sheet.cell(row=row, column=3).number_format = '$#,##0'
        row += 1
    money_row(row, "TOTAL GASTOS DEL DÍA", "", total_expenses, total=True)

    reconciliation_start = row + 2
    section("5. CONCILIACIÓN FINAL DE CAJA", reconciliation_start)
    money_row(reconciliation_start + 1, "Efectivo registrado en sistema", "Ventas en efectivo", totals["cash"])
    money_row(reconciliation_start + 2, "Gastos registrados del día", "Salidas de caja", total_expenses)
    money_row(
        reconciliation_start + 3,
        "EFECTIVO NETO ESPERADO EN CAJA",
        "Efectivo - gastos",
        totals["cash"] - total_expenses,
        total=True,
    )

    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 22

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"cierre-caja-{day.isoformat()}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _is_custom_acai(name: str) -> bool:
    normalized = "".join(c for c in unicodedata.normalize("NFKD", name) if not unicodedata.combining(c))
    return normalized.strip().casefold() == "acai personalizado"


def _acai_size(note: str | None) -> str | None:
    match = re.search(r"configurado:\s*vaso\s*(8|12|16)\s*oz\b", note or "", re.IGNORECASE)
    return {"8": "pequeño", "12": "mediano", "16": "grande"}.get(match.group(1)) if match else None


@router.get("/summary/products", response_model=list[schemas.SalesByProductOut])
def sales_by_product(period: str | None = None, db_session: Session = Depends(db.get_db)):
    start_date = _period_start(period)
    query = (
        db_session.query(
            models.Sale.order_id,
            models.SaleItem.menu_item_id,
            models.SaleItem.name,
            models.SaleItem.category,
            models.SaleItem.unit_price,
            models.SaleItem.quantity,
            models.SaleItem.line_total,
        )
        .join(models.Sale, models.Sale.id == models.SaleItem.sale_id)
    )
    notes_query = (
        db_session.query(models.PosOrderItem)
        .join(models.Sale, models.Sale.order_id == models.PosOrderItem.order_id)
    )
    if start_date is not None:
        query = query.filter(models.Sale.created_at >= start_date)
        notes_query = notes_query.filter(models.Sale.created_at >= start_date)

    # Historical sale lines do not retain the order-line ID or its note.
    # Match their stored attributes without joining and multiplying sale totals.
    sizes = defaultdict(set)
    for item in notes_query.all():
        if _is_custom_acai(item.name):
            key = (item.order_id, item.menu_item_id, item.unit_price, item.quantity)
            sizes[key].add(_acai_size(item.note))

    grouped = {}
    for row in query.all():
        name = row.name
        if _is_custom_acai(name):
            matches = sizes.get((row.order_id, row.menu_item_id, row.unit_price, row.quantity), set())
            size = next(iter(matches)) if len(matches) == 1 else None
            name = f"Açaí de vaso {size}" if size else "Açaí personalizado (sin tamaño registrado)"
        key = (row.menu_item_id, name, row.category)
        if key not in grouped:
            grouped[key] = schemas.SalesByProductOut(
                menu_item_id=row.menu_item_id, name=name, category=row.category,
                quantity=Decimal("0"), total=Decimal("0"),
            )
        grouped[key].quantity += row.quantity
        grouped[key].total += row.line_total
    return sorted(grouped.values(), key=lambda row: row.total, reverse=True)


@router.get("/summary/tables", response_model=list[schemas.SalesByTableOut])
def sales_by_table(period: str | None = None, db_session: Session = Depends(db.get_db)):
    start_date = _period_start(period)
    query = (
        db_session.query(
            models.PosOrder.table_id,
            models.PosTable.name,
            models.PosTable.is_active,
            func.coalesce(func.count(models.Sale.id), 0).label("quantity"),
            func.coalesce(func.sum(models.Sale.total), 0).label("total"),
        )
        .join(models.PosOrder, models.PosOrder.id == models.Sale.order_id)
        .outerjoin(models.PosTable, models.PosTable.id == models.PosOrder.table_id)
        .group_by(models.PosOrder.table_id, models.PosTable.name, models.PosTable.is_active)
        .order_by(func.sum(models.Sale.total).desc())
    )
    if start_date is not None:
        query = query.filter(models.Sale.created_at >= start_date)
    rows = query.all()
    return [
        schemas.SalesByTableOut(
            table_id=row.table_id,
            name=row.name,
            is_active=row.is_active,
            quantity=row.quantity,
            total=row.total,
        )
        for row in rows
    ]


@router.get("/summary/adjustments/monthly", response_model=list[schemas.SalesAdjustmentsByMonthOut])
def sales_adjustments_by_month(
    period: str | None = None,
    db_session: Session = Depends(db.get_db),
):
    start_date = _period_start(period)
    year_expr = func.extract("year", models.Sale.created_at)
    month_expr = func.extract("month", models.Sale.created_at)

    query = (
        db_session.query(
            year_expr.label("year"),
            month_expr.label("month"),
            func.coalesce(
                func.sum(case((models.PosOrderItem.courtesy.is_(True), 1), else_=0)),
                0,
            ).label("courtesy_count"),
            func.coalesce(
                func.sum(case((models.PosOrderItem.discount_amount > 0, 1), else_=0)),
                0,
            ).label("discount_count"),
        )
        .join(models.PosOrder, models.PosOrder.id == models.Sale.order_id)
        .outerjoin(models.PosOrderItem, models.PosOrderItem.order_id == models.PosOrder.id)
        .group_by(year_expr, month_expr)
        .order_by(year_expr.desc(), month_expr.desc())
    )
    if start_date is not None:
        query = query.filter(models.Sale.created_at >= start_date)
    rows = query.all()
    return [
        schemas.SalesAdjustmentsByMonthOut(
            year=int(row.year or 0),
            month=int(row.month or 0),
            courtesy_count=int(row.courtesy_count or 0),
            discount_count=int(row.discount_count or 0),
        )
        for row in rows
    ]
